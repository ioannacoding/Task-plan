from tkinter import*
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

root=Tk()
root.title("PlanTasks")
root.iconbitmap('icon.ico')
root.geometry('500x800')

wb=Workbook()

wb=load_workbook('tasks.xlsx')
ws=wb.active

column_a=ws['A']
column_b=ws['B']
column_c=ws['C']

print(column_a)

def get_a():
    list=''
    for i in column_a:
        list=f'{list +str(i.value)}\n'
    label_a.config(text=list)

def get_b():
    list=''
    for cell in column_b:
        print(cell.value)
        list=f'{list +str(cell.value)}\n'
    label_b.config(text=list)

def get_c():
    list=''
    for cell in column_c:
        print(cell.value)
        list=f'{list +str(cell.value)}\n'
    label_c.config(text=list)

ba=Button(root,text="Get Column A", command=get_a)
ba.pack(pady=20)

bb=Button(root,text="Get Column B", command=get_b)
bb.pack(pady=20)

bc=Button(root,text="Get Column C", command=get_c)
bc.pack(pady=20)



label_a=Label(root,text="good morning")
label_a.pack(pady=20)

label_b=Label(root,text="good night")
label_b.pack(pady=20)

label_c=Label(root,text="good afternoon")
label_c.pack(pady=20)

ws['A8']="Eat Cheese"
ws['B8']="Said Hello"

wb.save('tasks.xlsx')

root.mainloop()

